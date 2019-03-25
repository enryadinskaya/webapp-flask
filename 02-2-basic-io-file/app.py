import os

from flask import Flask, render_template, request, flash, redirect, url_for, send_from_directory

from wtforms import Form, TextField, validators

from openpyxl import load_workbook

from fpdf import FPDF

from werkzeug.utils import secure_filename


UPLOAD_FOLDER = ''
ALLOWED_EXTENSIONS = set(['txt', 'pdf', 'png', 'jpg', 'jpeg', 'gif','xlsx','xls'])
app_new = Flask(__name__, static_folder='')
app_new.secret_key = 'another_secret'
app_new.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER


class ReusableForm(Form):
    file = TextField('File:', validators=[validators.required()])


def massiv(file):
    wb = load_workbook(file)
    mas = list()

    sheetname = wb.get_sheet_names()

    sheet = wb.get_sheet_by_name(sheetname[0])

    massiv = list()

    for cellObj in sheet['A1':'T20']:
        string = []
        for cell in cellObj:
            if cell.value == None:
                break
            else:
                string.append(cell.value)
        if string != []:
            massiv.append(string)

    line = len(massiv)
    column = len(massiv[0])

    newmassiv = [[0]*line for i in range (column)]

    for i in range(line):
        for j in range(column):
            newmassiv[j][i] = massiv[i][j]

    for i in range(column):
        newline = ''
        for j in range(line):
            newline = newline + '\t' + str(newmassiv[i][j])
        mas.append(newline)

    pdf = FPDF()
    pdf.add_page()
    pdf.set_xy(0, 0)
    pdf.set_font('arial', 'B', 13.0)
    for i in range(len(mas)):
        pdf.ln(5)
        pdf.cell(h=5.0, align='L', w=0, txt=mas[i])
        pdf.ln(1)
    filepdf = 'test.pdf'
    pdf.output(filepdf, 'F')

    return filepdf

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1] in ALLOWED_EXTENSIONS

@app_new.route("/", methods=['GET', 'POST'])
def mainPage():

    form = ReusableForm(request.form)

    if request.method == 'POST':
        file = request.files['file']
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            file.save(os.path.join(app_new.config['UPLOAD_FOLDER'], filename))
            filedoc = massiv(filename)
            return redirect(url_for('uploaded_file', filedoc=filedoc))

    else:
        flash('Error: All the form fields are required. ')

    return render_template('index.html', form=form)


@app_new.route('/<path:filedoc>')
def uploaded_file(filedoc):
    return send_from_directory(app_new.static_folder,
                               filedoc)

if __name__ == '__main__':
    app_new.run(debug=True, host='0.0.0.0', port=5003)
